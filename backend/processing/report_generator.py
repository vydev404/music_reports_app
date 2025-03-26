# -*- coding: utf-8 -*-
from pathlib import Path

import openpyxl
from processing.config import settings
from processing.schemas.dto import (
    FilteredClipDTO,
    ReportDTO,
    SourceFileDTO,
    UsedMusicDTO,
)
from utils.timecode_converter import TimecodeConverter as tc


class ReportsGenerator:
    def __init__(
        self,
        report_template: str = settings.report.template,
        reports_dir: str = settings.report.output_dir,
    ) -> None:
        self.report_template = Path(report_template)
        self.reports_dir = Path(reports_dir)

    def generate_report(self, source_file: SourceFileDTO) -> ReportDTO:
        report_file = self.reports_dir / f"{source_file.name}.xlsx"
        report_file = self.rename_if_exists(report_file)
        report = ReportDTO(
            name=source_file.name,
            source_file_id=source_file.id,
            file=str(report_file),
        )

        header_data = {
            "C3": source_file.name,
            "C4": source_file.created,
            "C5": source_file.data.total_duration["timecode"],
            "C6": 0,
        }

        wb = openpyxl.load_workbook(self.report_template)

        ws = wb.active

        clips = source_file.data.items.get("items_in_db", [])

        for row_num, clip in enumerate(clips, start=10):
            if not isinstance(clip, FilteredClipDTO):
                continue
            music_item = UsedMusicDTO(
                id=clip.clip_info.id,
                title=clip.clip_info.title,
                count=clip.count,
                duration=clip.duration["timecode"],
            ).to_dict()
            report.used_music.append(music_item)
            header_data["C6"] += clip.duration["frames"]
            ws.cell(row=row_num, column=1, value=row_num - 9)
            ws.cell(
                row=row_num,
                column=2,
                value=clip.clip_info.title if clip.clip_info else "N/A",
            )
            ws.cell(
                row=row_num,
                column=3,
                value=f"{clip.clip_info.artist if clip.clip_info else ""}",
            )
            ws.cell(
                row=row_num,
                column=4,
                value=clip.clip_info.album if clip.clip_info else "",
            )
            ws.cell(row=row_num, column=5, value=clip.duration["timecode"])
            ws.cell(row=row_num, column=6, value=source_file.name)
            ws.cell(
                row=row_num,
                column=7,
                value=clip.clip_info.right_holder if clip.clip_info else "N/A",
            )

        header_data["C6"] = tc.frames_to_timecode(header_data["C6"])
        for k, v in header_data.items():
            ws[k].value = v
        wb.save(report.file)

        return report

    def generate_not_in_db_list(self, source_file: SourceFileDTO):
        clips = source_file.data.items.get("items_not_in_db", [])
        not_included = self.reports_dir / f"{source_file.name}_not_in_db.txt"
        if not clips:
            return
        with open(not_included, "w", encoding="utf-8") as f:
            header = f"[*]  PARSED CLIP FILES FROM {source_file.name} NOT FOUNDED IN DATABASE\n\n"
            f.write(header)
            for i, clip in enumerate(clips, start=1):
                f.write(
                    f"[{i:02}] {clip.clip_name: ^70}| used: {clip.count: ^3} times | total duration: {clip.duration["timecode"]: ^9}\n"
                )

    @staticmethod
    def rename_if_exists(file: Path) -> Path:
        count = 1
        tmp = file
        while tmp.exists():
            tmp = f"{file.stem}_{count:02}{file.suffix}"
            count += 1
        return tmp
